import pandas as pd
import io

def export_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="QA Results")

        wb = writer.book
        ws = writer.sheets["QA Results"]

        from openpyxl.styles import PatternFill, Font
        SEV_FILLS = {
            "Critical": PatternFill("solid", fgColor="FCEBEB"),
            "Major":    PatternFill("solid", fgColor="FAECE7"),
            "Minor":    PatternFill("solid", fgColor="FAEEDA"),
            "Pass":     PatternFill("solid", fgColor="EAF3DE"),
        }
        SEV_COL = None
        for i, col in enumerate(df.columns, 1):
            if col == "severity":
                SEV_COL = i
                break

        if SEV_COL:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                sev_cell = row[SEV_COL - 1]
                fill = SEV_FILLS.get(sev_cell.value)
                if fill:
                    for cell in row:
                        cell.fill = fill

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    return buf.getvalue()


def export_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")
